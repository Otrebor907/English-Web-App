"""Allinea un workbook del programma allo schema Neon attuale.

Uso: python3 aggiorna_workbook.py <file.xlsx> [altro.xlsx ...]

Idempotente: rieseguirlo su un file gia' aggiornato non cambia nulla.

Fa due cose:
1. ripulisce le colonne e i sottotitoli dei fogli letti dall'importer, quando
   restano tracce di campi rimossi dal modello dati;
2. elimina i fogli di documentazione che nessuna parte del codice legge.

Il punto 2 e' un'inversione rispetto alla versione precedente di questo script,
che quei fogli li *creava*. Erano una copia a mano dello schema del database,
da riallineare a ogni migration: informazione duplicata che invecchiava in
silenzio (il foglio «Modello Dati» dichiarava ancora i nomi mapping_* dopo che
la migration 0008 li aveva rinominati in dim_*). La fonte di verita' dello
schema e' backend/learning/models.py, dove i nomi delle tabelle sono espliciti
in db_table. Il workbook resta la fonte del *programma didattico*, non dello
schema. Fogli eliminati il 31/08/2026; il contenuto dei due fogli che non
duplicavano il codice e' conservato in Doc/Logica_Didattica.md.

--- Nota storica: corrispondenza dei nomi tabella prima e dopo il refactor ---
  dim_area_lezione          <- mapping_area_lezione <- learning_area
  dim_tipologia             <- mapping_tipologia    <- learning_tipologia
  dim_livello               <- mapping_livello      <- learning_livello
  dim_difficolta_lezione    <- mapping_difficolta_lezione <- learning_difficolta
  dim_stato_lezione         <- learning_statolezione
  struttura_lezione         <- learning_sezionelezione
  struttura_quiz            <- learning_quiz
  struttura_quiz_guidato    <- learning_quesito (parte guidata)
  struttura_quiz_finale     <- learning_quesito (parte finale)
  user_progress             <- learning_progresso
  user_profile              <- learning_user
  user_authtoken_token      <- authtoken_token
  learning_importanza       ELIMINATA, con la colonna learning_lezione.priorita
  learning_prerequisito     ELIMINATA (vedi sotto)

--- Nota storica: perche' learning_prerequisito e' stata eliminata ---
I 119 archi prerequisito non bloccavano nulla: le lezioni sono sempre state
tutte accessibili, e il grafo serviva solo a mostrare un consiglio («Segue X»,
«prerequisiti mancanti»). A fronte di questo, costava una tabella ponte, una
M2M Lezione<->Lezione, la validazione del DAG in services.py e una colonna nel
workbook. Rimosso tutto: l'ordine con cui affrontare le lezioni resta espresso
da ordine_mvp (1..29 sul percorso MVP) e da ordine_percorso (1..98 sul
programma completo). Se un domani servira' un vero blocco per prerequisiti, si
reintroduce allora, su un prodotto gia' avviato.
"""
import sys
import warnings

warnings.filterwarnings("ignore")

from openpyxl import load_workbook

# I soli fogli che l'importer legge davvero (learning/importer.py, load_source):
# vanno preservati sempre. Ogni altro foglio e' documentazione o residuo.
FOGLI_LETTI = {
    "Programma Lezioni", "Percorso MVP", "Grammatica",
    "Vocabolario", "Comunicazione", "Liste",
}

# Fogli da rimuovere. I primi cinque erano una copia dello schema del database;
# gli ultimi due contenevano il ragionamento didattico e la roadmap di prodotto,
# ora in Doc/Logica_Didattica.md.
FOGLI_OBSOLETI = [
    "Modello Dati",
    "Struttura_Lezione",
    "Struttura_Quiz",
    "struttura_quiz_guidato",
    "struttura_quiz_finale",
    "Logica Didattica",
    "Roadmap",
    "Pronuncia",  # solo nella vecchia versione con audio del workbook
]


def _elimina_colonna(ws, intestazione, riga_intestazioni=4):
    valori = [c.value for c in ws[riga_intestazioni]]
    if intestazione in valori:
        ws.delete_cols(valori.index(intestazione) + 1)
        return True
    return False


def aggiorna(percorso):
    wb = load_workbook(percorso)
    modifiche = []

    if "Liste" in wb.sheetnames:
        ws = wb["Liste"]
        if _elimina_colonna(ws, "Importanza MVP"):
            modifiche.append("Liste: rimossa colonna «Importanza MVP»")
        ws["A2"] = (
            "Queste colonne alimentano le convalide dati degli altri fogli. Ogni colonna corrisponde a una tabella "
            "su Neon: Area Didattica → dim_area_lezione, Tipologia Lezione → dim_tipologia, "
            "Livello Linguistico → dim_livello, Difficoltà → dim_difficolta_lezione, "
            "Stato della Lezione → dim_stato_lezione."
        )

    if "Percorso MVP" in wb.sheetnames:
        ws = wb["Percorso MVP"]
        if _elimina_colonna(ws, "Importanza"):
            modifiche.append("Percorso MVP: rimossa colonna «Importanza»")
        for cella in ws[4]:
            if cella.value == "Motivazione della Priorità":
                cella.value = "Motivazione della Scelta"
                modifiche.append("Percorso MVP: «Motivazione della Priorità» → «Motivazione della Scelta»")
        ws["A2"] = (
            "Le lezioni del percorso MVP. L'ordine è dato da «Ordine MVP» (learning_lezione.ordine_mvp): la "
            "classificazione Essenziale/Consigliata/Secondaria è stata rimossa dal modello dati insieme alla "
            "tabella learning_importanza."
        )

    for nome in FOGLI_OBSOLETI:
        # Cintura di sicurezza: un foglio letto dall'importer non si tocca mai,
        # qualunque cosa dica la lista qui sopra.
        if nome in wb.sheetnames and nome not in FOGLI_LETTI:
            del wb[nome]
            modifiche.append(f"Rimosso foglio «{nome}» (non letto da nessuna parte del codice)")

    mancanti = FOGLI_LETTI - set(wb.sheetnames)
    if mancanti:
        raise SystemExit(f"ERRORE: {percorso} non ha i fogli richiesti dall'importer: {', '.join(sorted(mancanti))}")

    wb.save(percorso)
    return modifiche or ["Nessuna modifica: il file era già allineato"]


if __name__ == "__main__":
    for percorso in sys.argv[1:]:
        print(f"\n{percorso}")
        for voce in aggiorna(percorso):
            print(f"  - {voce}")
