import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from django.db import transaction
from openpyxl import load_workbook
from .models import (
    Area, Difficolta, Lezione, Livello, QuesitoFinale, QuesitoGuidato,
    StatoLezione, StrutturaLezione, StrutturaQuiz, Tipologia,
)


LOOKUP_MODELS = {
    "area": Area,
    "tipologia": Tipologia,
    "livello": Livello,
    "difficolta": Difficolta,
    "stato": StatoLezione,
}
# Aree foglio Excel
SUPPORTED_AREAS = {"GRA", "VOC", "COM"}
# Oridne e numero quantità sezioni per area, come da fogli Excel Grammatica, Vocabolario e Comunicazione.
AREA_SECTION_COUNTS = {"GRA": 9, "VOC": 7, "COM": 8}
# Ttoale lezioni in Programma Lezioni e lezioni MVP previste dal programma (foglio Percorso MVP).
CATALOG_LESSON_COUNT = 98
MVP_LESSON_COUNT = 29
# Colonne obbligatorie per ogni lezione, come da foglio Programma Lezioni
REQUIRED_LESSON_FIELDS = {
    "id", "area", "tipologia", "nome", "livello", "difficolta", "ordine_percorso",
    "obiettivo_didattico", "competenze", "durata_min", "errori_tipici", "stato",
}
REQUIRED_COMPLETE_SOURCE_KEYS = {"liste", "lezioni", "sezioni", "quiz"}


def _sheet_rows_at(sheet, header_row, max_column=None, max_row=None):
    rows = sheet.iter_rows(min_row=header_row, max_row=max_row, values_only=True, max_col=max_column)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    return [
        {headers[index]: value for index, value in enumerate(row) if headers[index]}
        for row in rows if any(value is not None for value in row)
    ]
# Funzione simile a _sheet_rows ma permette di specificare l'indice della riga di intestazione e limiti di colonna/riga.
# esempio output:
# righe[2]  →  riga 7 del foglio  →  GRA-A1-003
#   {
#       'ID Lezione': 'GRA-A1-003',
#       'Area Didattica': 'Grammatica'ecc
#   }

def _code(value):
    normalized = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]+", "_", normalized.upper()).strip("_")


def _split_list(value):
    if value in (None, "", "—") or str(value).strip().casefold() in {"nessuno", "nessuna"}:
        return []
    return [item.strip() for item in str(value).split(",") if item.strip()]


def _with_expected_counts(data, fixture=False):
    meta = data.setdefault("meta", {})
    meta.setdefault("catalogo_atteso", len(data.get("lezioni", [])) if fixture else CATALOG_LESSON_COUNT)
    meta.setdefault(
        "mvp_atteso",
        sum(row.get("ordine_mvp") is not None for row in data.get("lezioni", [])) if fixture else MVP_LESSON_COUNT,
    )
    return data


def _audio_field_paths(value, path="fonte"):
    paths = []
    if isinstance(value, dict):
        for key, child in value.items():
            child_path = f"{path}.{key}"
            if "audio" in str(key).casefold():
                paths.append(child_path)
            paths.extend(_audio_field_paths(child, child_path))
    elif isinstance(value, list):
        for index, child in enumerate(value):
            paths.extend(_audio_field_paths(child, f"{path}[{index}]"))
    return paths


def _reject_incomplete_json_fragment(data):
    if not isinstance(data, dict) or "id" not in data or "lezioni" in data:
        return
    missing_lesson_fields = sorted(
        field for field in REQUIRED_LESSON_FIELDS if data.get(field) in (None, "")
    )
    missing_blocks = sorted(REQUIRED_COMPLETE_SOURCE_KEYS)
    details = []
    if missing_lesson_fields:
        details.append(f"campi lezione mancanti: {', '.join(missing_lesson_fields)}")
    details.append(f"blocchi catalogo mancanti: {', '.join(missing_blocks)}")
    raise ValueError(
        f"Frammento JSON della lezione {data['id']} riconosciuto ma non importabile come catalogo completo; "
        + "; ".join(details)
    )


def _load_programma_workbook(workbook):
    """Map the supplied planning workbook without treating planned copy as finished content."""
    area_codes = {}
    for row in workbook["Liste"].iter_rows(min_row=16, max_row=18, min_col=1, max_col=2, values_only=True):
        if row[0] and row[1]:
            area_codes[str(row[0])] = str(row[1])
    # Produce il dizionario
    #   area_codes = {
    #       "Grammatica":    "GRA",
    #       "Vocabolario":   "VOC",
    #       "Comunicazione": "COM",
    #                  }
    
    list_columns = {
        "Area Didattica": "area", "Tipologia Lezione": "tipologia",
        "Livello Linguistico": "livello", "Difficoltà": "difficolta",
        "Stato della Lezione": "stato",
    }
    lists = defaultdict(list)
    seen = defaultdict(set)
    for row in _sheet_rows_at(workbook["Liste"], 4, 8, max_row=10):
        for source_name, target_name in list_columns.items():
            value = row.get(source_name)
            if not value:
                continue
            code = area_codes.get(str(value), str(value)) if target_name == "area" else (
                str(value) if target_name in {"livello", "difficolta"} else _code(value)
            )
            if code not in seen[target_name]:
                lists[target_name].append({"code": code, "nome": str(value)})
                seen[target_name].add(code)

    # Produce le 5 tabelle di DIM_TABLE che conosciamo
    #lists["area"] = [
    #    {'code': 'GRA', 'nome': 'Grammatica'},
    #    {'code': 'VOC', 'nome': 'Vocabolario'},
    #    {'code': 'COM', 'nome': 'Comunicazione'},
    #]
    #lists["tipologia"] = [
    #    {'code': 'REGOLA_ED_ESERCIZI', 'nome': 'Regola ed esercizi'},
    #    {'code': 'LESSICO_TEMATICO',   'nome': 'Lessico tematico'},
    #    {'code': 'DIALOGO',            'nome': 'Dialogo'},
    #    {'code': 'PRODUZIONE_GUIDATA', 'nome': 'Produzione Guidata'},
    #    {'code': 'PRODUZIONE_LIBERA',  'nome': 'Produzione Libera'},
    #]
    #lists["livello"] = [
    #    {'code': 'A1', 'nome': 'A1'}, {'code': 'A2', 'nome': 'A2'}, {'code': 'B1', 'nome': 'B1'},
    #    {'code': 'B2', 'nome': 'B2'}, {'code': 'C1', 'nome': 'C1'}, {'code': 'C2', 'nome': 'C2'},
    #]
    #lists["difficolta"] = [
    #    {'code': 'Bassa', 'nome': 'Bassa'}, {'code': 'Media', 'nome': 'Media'}, {'code': 'Alta', 'nome': 'Alta'},
    #]
    #lists["stato"] = [
    #    {'code': 'DA_SVILUPPARE',     'nome': 'Da sviluppare'},
    #    {'code': 'DA_SVILUPPARE_MVP', 'nome': 'Da sviluppare (MVP)'},
    #    {'code': 'IN_SVILUPPO',       'nome': 'In sviluppo'},
    #    {'code': 'IN_REVISIONE',      'nome': 'In revisione'},
    #    {'code': 'COMPLETATA',        'nome': 'Completata'},
    #    {'code': 'PUBBLICATA',        'nome': 'Pubblicata'},
    #]


    mvp = {}
    for row in _sheet_rows_at(workbook["Percorso MVP"], 4, 9):
        lesson_id = row.get("ID Lezione")
        if not lesson_id or not isinstance(row.get("Ordine MVP"), (int, float)):
            continue
        mvp[str(lesson_id)] = {"ordine_mvp": int(row["Ordine MVP"])}
    # Viene prodotto questa mappa, che è la sequenza di lezioni MVP prevista dal programma:
    #    mvp = {
    #   'GRA-A1-001': {'ordine_mvp': 1},
    #   'GRA-A1-003': {'ordine_mvp': 3},      # ← la nostra
    #   'GRA-A1-002': {'ordine_mvp': 2},
    #   'GRA-A1-004': {'ordine_mvp': 4},
    #   'COM-A1-001': {'ordine_mvp': 5},
    #   'VOC-A1-001': {'ordine_mvp': 6},
    #   'GRA-A1-005': {'ordine_mvp': 7},
    #   'GRA-A1-006': {'ordine_mvp': 8},
    #   'VOC-A1-002': {'ordine_mvp': 9},
    #   'GRA-A1-007': {'ordine_mvp': 10},
    #   'GRA-A1-008': {'ordine_mvp': 11},
    #   'GRA-A1-009': {'ordine_mvp': 12},
    #   'VOC-A1-003': {'ordine_mvp': 13},
    #   ...                                   # fino a 29
    #   }

    # Creata lista di lezioni con campi normalizzati e mappati, senza contenuti definitivi. (4 è il rigo di intestazione, 17 è l'ultima colonna utile)
    lessons = []
    for row in _sheet_rows_at(workbook["Programma Lezioni"], 4, 17):
        lesson_id = row.get("ID Lezione")
        if not lesson_id:
            continue
        lesson_id = str(lesson_id)
        mvp_row = mvp.get(lesson_id)
        lessons.append({
            "id": lesson_id,
            "area": area_codes[str(row["Area Didattica"])],
            "tipologia": _code(row["Tipologia Lezione"]),
            "nome": str(row["Nome Lezione"]),
            "descrizione": str(row.get("Breve Descrizione") or ""),
            "categoria": str(row.get("Categoria") or "").strip(),
            "livello": str(row["Livello Linguistico"]),
            "difficolta": str(row["Difficoltà"]),
            "ordine_percorso": int(row["Ordine nel Percorso"]),
            "obiettivo_didattico": str(row["Obiettivo Didattico"]),
            "competenze": _split_list(row.get("Competenze Allenate")), #applicata funzione di split per trasformare esempio stringa "Grammatica, Lettura" in lista ["Grammatica", "Lettura"]
            "durata_min": int(row["Durata Stimata (min)"]),
            "errori_tipici": [str(row["Errori Tipici degli Italiani"])],
            "stato": _code(row["Stato della Lezione"]),
            "ordine_mvp": mvp_row["ordine_mvp"] if mvp_row else None,
        })
    # Restituisce in dizionario il catalogo completo di lezioni, con campi normalizzati e mappati, senza contenuti definitivi.
    #   Lessons[2] = {
    #       id': 'GRA-A1-003',
    #       'area': 'GRA',                                     # ← area_codes['Grammatica']
    #       'tipologia': 'REGOLA_ED_ESERCIZI',                 # ← _code('Regola ed esercizi')
    #       'nome': 'Il verbo to be: forma affermativa',
    #       'descrizione': 'Regola spiegata in parole semplici, strutture (affermativa, negativa, '
    #                      "interrogativa), esempi con traduzione, confronto con l'italiano ed "
    #                      'esercizi progressivi.',
    #       'categoria': 'Il verbo to be',
    #       'livello': 'A1',
    #       'difficolta': 'Bassa',
    #       'ordine_percorso': 3,
    #       'obiettivo_didattico': 'Coniugare to be al presente e usarlo per età, nazionalità, '
    #                              'professione e stati.',
    #       'competenze': ['Grammatica', 'Lettura'],           # ← _split_list: str → list
    #       'durata_min': 12,
    #       'errori_tipici': ['«I have 25 years» invece di «I am 25»; '
    #                         '«I am agree» invece di «I agree».'],   # ← avvolta in una lista
    #       'stato': 'DA_SVILUPPARE_MVP',                      # ← _code('Da sviluppare (MVP)')
    #       'ordine_mvp': 3,                                   # ← NON dalla riga: da mvp['GRA-A1-003']
    #       }
    #       **tre campi spariscono**: `Prerequisiti`, `Lezione Precedente`, `Lezione Successiva`.

    template_sheets = {"GRA": "Grammatica", "VOC": "Vocabolario", "COM": "Comunicazione"}
    templates = {}
    for area, sheet_name in template_sheets.items():
        templates[area] = []
        for row in _sheet_rows_at(workbook[sheet_name], 4, 5):
            if not isinstance(row.get("Ordine"), (int, float)):
                break
            templates[area].append({
                "ordine": int(row["Ordine"]),
                "tipo_sezione": str(row["Sezione della Lezione"]),
                "contenuto": {
                    "titolo": str(row["Sezione della Lezione"]),
                    "todo": f"TODO_FONTE: {row['Contenuto Previsto']}",
                    "obiettivo_template": str(row["Obiettivo"]),
                    "formato_sorgente": str(row["Formato Web App"]),
                },
                "formato_web": "todo",
            })
    sections = [
        {"lezione_id": lesson["id"], **section}
        for lesson in lessons for section in templates[lesson["area"]]
    ]
    return {
        "liste": dict(lists), "lezioni": lessons,
        "sezioni": sections, "quiz": [],
        "meta": {
            "formato": "programma_lezioni",
            "catalogo_atteso": CATALOG_LESSON_COUNT,
            "mvp_atteso": MVP_LESSON_COUNT,
            "quiz_mancanti": True,
            "contenuti_sezioni_todo": True,
        },
    }
    # Fa sostalzialmente per ogni lezione scritta negli sheet grammatica - vocabolario - comunicazione, una copia dello scheletro delle sezioni previste (prodotto cartesioni con le parti scritte nella colonna A)
    # Risultato atteso ad esempio per solo due punti su 9 previsti 
    #{
    #'ordine': 1,
    #'lezione_id': 'GRA-A1-003',
    #'tipo_sezione': 'Obiettivo della lezione',
    #'contenuto': {
    #    'titolo': 'Obiettivo della lezione',
    #            'in questo momento»).',
    #    'todo': 'TODO_FONTE: Una frase concreta («Alla fine saprai dire cosa stai facendo '
    #    'obiettivo_template': 'Dire subito cosa si saprà fare alla fine.',
    #    'formato_sorgente': 'Banner in evidenza',
    #},
    #'formato_web': 'todo',
    #}
    #La quinta, quella degli errori tipici:
    #    'lezione_id': 'GRA-A1-003',
    #{
    #    'ordine': 5,
    #    'tipo_sezione': 'Errori tipici degli italiani',
    #    'contenuto': {
    #        'titolo': 'Errori tipici degli italiani',
    #        'todo': 'TODO_FONTE: Frase sbagliata → frase corretta → perché '
    #                '(es. «I have 25 years» → «I am 25»).',
    #    'obiettivo_template': "Prevenire i calchi dall'italiano.",
    #    'formato_sorgente': 'Box confronto ❌/✅',
    #    },
    #}
    #    'formato_web': 'todo',

# Le due sole fonti importabili sono il workbook del programma e il JSON.
# Il workbook porta il catalogo e lo scheletro delle pagine (sezioni TODO_FONTE,
# nessun quiz); il JSON serve alle fixture dei test, che hanno invece contenuti
# e quesiti completi. I testi definitivi delle lezioni vere non passano da qui:
# arrivano dai brief markdown, con il comando pubblica_da_markdown
# (learning/markdown_source.py).
NATIVE_SHEETS = {"Programma Lezioni", "Percorso MVP", "Grammatica", "Vocabolario", "Comunicazione", "Liste"}


def load_source(path):
    path = Path(path)
    if path.suffix.lower() == ".json":
        with path.open(encoding="utf-8") as source:
            data = json.load(source)
        _reject_incomplete_json_fragment(data)
        return _with_expected_counts(data, fixture=data.get("meta", {}).get("tipo") == "fixture_tecnica")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Formato non supportato: usare .json, .xlsx o .xlsm")
    workbook = load_workbook(path, data_only=True, read_only=True)
    missing = NATIVE_SHEETS - set(workbook.sheetnames)
    if missing:
        raise ValueError(
            f"Fogli Excel mancanti: {', '.join(sorted(missing))}. "
            f"Il file deve essere il workbook del programma, con i fogli: {', '.join(sorted(NATIVE_SHEETS))}"
        )
    return _load_programma_workbook(workbook)


def collect_source_structure_errors(data):
    errors = []
    audio_fields = _audio_field_paths(data)
    if audio_fields:
        errors.append(f"Campi audio non ammessi: {', '.join(audio_fields)}")
    for key in LOOKUP_MODELS:
        if key not in data.get("liste", {}):
            errors.append(f"Lista lookup mancante: {key}")
    lessons = data.get("lezioni", [])
    if not lessons:
        errors.append("Nessuna lezione nel file sorgente")
    lookup_codes = {
        key: {str(item["code"]) for item in data.get("liste", {}).get(key, [])}
        for key in LOOKUP_MODELS
    }
    if lookup_codes["area"] != SUPPORTED_AREAS:
        errors.append(
            "Le aree della fonte devono essere esclusivamente GRA, VOC e COM; "
            f"trovate: {', '.join(sorted(lookup_codes['area'])) or 'nessuna'}"
        )
    lookup_fields = {
        "area": "area", "tipologia": "tipologia", "livello": "livello",
        "difficolta": "difficolta", "stato": "stato",
    }
    orders = []
    mvp_orders = []
    for lesson in lessons:
        missing_fields = sorted(key for key in REQUIRED_LESSON_FIELDS if lesson.get(key) in (None, ""))
        if missing_fields:
            errors.append(f"{lesson.get('id', '<senza id>')}: campi obbligatori mancanti: {', '.join(missing_fields)}")
        for key, field in lookup_fields.items():
            value = lesson.get(field)
            if value is not None and str(value) not in lookup_codes[key]:
                errors.append(f"{lesson.get('id')}: valore {field} inesistente: {value}")
        if lesson.get("area") not in AREA_SECTION_COUNTS:
            errors.append(f"{lesson.get('id')}: area non supportata")
        if isinstance(lesson.get("ordine_percorso"), int):
            orders.append(lesson["ordine_percorso"])
            if not 1 <= lesson["ordine_percorso"] <= CATALOG_LESSON_COUNT:
                errors.append(f"{lesson.get('id')}: ordine_percorso fuori intervallo 1..98")
        if lesson.get("ordine_mvp") is not None:
            mvp_orders.append(lesson["ordine_mvp"])

    if sorted(orders) != list(range(1, len(lessons) + 1)):
        errors.append(f"ordine_percorso deve essere una sequenza unica 1..{len(lessons)}")
    if sorted(mvp_orders) != list(range(1, len(mvp_orders) + 1)):
        errors.append(f"ordine_mvp deve essere una sequenza unica 1..{len(mvp_orders)}")
    expected_lessons = data.get("meta", {}).get("catalogo_atteso")
    expected_mvp = data.get("meta", {}).get("mvp_atteso")
    if expected_lessons is not None and len(lessons) != expected_lessons:
        errors.append(f"Catalogo incompleto: attese {expected_lessons} lezioni, trovate {len(lessons)}")
    if expected_mvp is not None and len(mvp_orders) != expected_mvp:
        errors.append(f"Percorso MVP incompleto: attese {expected_mvp} lezioni, trovate {len(mvp_orders)}")

    lesson_ids = {row["id"] for row in lessons}
    section_counts = defaultdict(int)
    for section in data.get("sezioni", []):
        if section["lezione_id"] not in lesson_ids:
            errors.append(f"Sezione riferita a lezione inesistente: {section['lezione_id']}")
        section_counts[section["lezione_id"]] += 1
    for lesson in lessons:
        expected = AREA_SECTION_COUNTS.get(lesson.get("area"))
        if expected is None:
            continue
        if section_counts[lesson["id"]] != expected:
            errors.append(f"{lesson['id']}: attese {expected} sezioni, trovate {section_counts[lesson['id']]}")

    for quiz in data.get("quiz", []):
        if quiz["lezione_id"] not in lesson_ids:
            errors.append(f"Quiz riferito a lezione inesistente: {quiz['lezione_id']}")
        questions = quiz.get("quesiti", [])
        if quiz["modalita"] == StrutturaQuiz.FINALE and not 8 <= len(questions) <= 10:
            errors.append(f"{quiz['lezione_id']}: il quiz finale deve avere 8-10 quesiti")
        for question in questions:
            if question["tipo"] not in {QuesitoFinale.SCELTA_MULTIPLA, QuesitoFinale.COMPLETAMENTO}:
                errors.append(f"Tipo quesito non valido: {question['tipo']}")
            if question["tipo"] == QuesitoFinale.SCELTA_MULTIPLA and question["risposta_corretta"] not in question.get("opzioni", []):
                errors.append("La risposta corretta deve essere presente nelle opzioni")
    return errors

    ### I controlli che vengono eseguiti da `collect_source_structure_errors` sono riassunti nella tabella seguente, con i risultati per il workbook GRA-A1-003.
    #| --- | --- |
    #| Controllo | Per GRA-A1-003 / per il file |
    #| Nessun campo il cui nome contenga «audio», a qualsiasi profondità | ✅ nessuno |
    #| Le aree devono essere esattamente `{GRA, VOC, COM}` | ✅ |
    #| Ogni lezione ha tutti i campi di `REQUIRED_LESSON_FIELDS` | ✅ 12 campi presenti |
    #| I valori di lookup esistono nelle liste | ✅ `GRA`, `REGOLA_ED_ESERCIZI`, `A1`, `Bassa`, `DA_SVILUPPARE_MVP` |
    #| `ordine_percorso` è una permutazione esatta di 1..98 | ✅ |
    #| `ordine_mvp` è una permutazione esatta di 1..29 | ✅ (grazie alla normalizzazione) |
    #| Ogni lezione ha il numero di sezioni della sua area | ✅ 9 per GRA-A1-003 |
    #| Il quiz finale, se presente, ha 8–10 quesiti | — nessun quiz dal workbook |


def validate_source(data):
    structure_errors = collect_source_structure_errors(data)
    if structure_errors:
        raise ValueError("Fonte non valida:\n- " + "\n- ".join(structure_errors))


def source_report(data):
    lessons = data.get("lezioni", [])
    mvp_ids = {row["id"] for row in lessons if row.get("ordine_mvp") is not None}
    errors = collect_source_structure_errors(data)
    warnings = []
    published_count = sum(row.get("stato") == "PUBBLICATA" and row.get("ordine_mvp") is not None for row in lessons)
    if not published_count:
        warnings.append("Nessuna lezione MVP ha stato PUBBLICATA")
    if not data.get("quiz"):
        warnings.append("Il file sorgente non contiene quesiti o quiz importabili")
    todo_sections = sum(str(section.get("contenuto", {}).get("todo", "")).startswith("TODO_FONTE:") for section in data.get("sezioni", []))
    if todo_sections:
        warnings.append(f"{todo_sections} sezioni attendono contenuti definitivi (TODO_FONTE)")
    return {
        "valido": not errors,
        "formato": data.get("meta", {}).get("formato", "normalizzato"),
        "conteggi": {
            "lezioni": len(lessons), "lezioni_mvp": len(mvp_ids), "lezioni_mvp_pubblicate": published_count,
            "sezioni": len(data.get("sezioni", [])),
            "sezioni_todo": todo_sections, "quiz": len(data.get("quiz", [])),
        },
        "errori": errors,
        "avvisi": warnings,
    }


@transaction.atomic
def import_content(path):
    # Produce il dizionario
    data = load_source(path)
    # controlla o ferma il ritmo se ci sono errori di struttura
    validate_source(data)

    ## SQL insert/update delle Dimension Tables (lookup)
    for key, model in LOOKUP_MODELS.items():
        source_codes = []
        for item in data["liste"][key]:
            code = str(item["code"])
            source_codes.append(code)
            model.objects.update_or_create(code=code, defaults={"nome": item["nome"]})

    ## SQL insert/update delle lezioni e delle sezioni
    source_ids = []
    for row in data["lezioni"]:
        source_ids.append(row["id"])
        defaults = {key: row.get(key, "") for key in (
            "nome", "descrizione", "categoria", "ordine_percorso", "obiettivo_didattico", "competenze",
            "durata_min", "errori_tipici", "ordine_mvp",
        )}
        defaults.update({
            "area_id": row["area"], "tipologia_id": row["tipologia"], "livello_id": row["livello"],
            "difficolta_id": row["difficolta"], "stato_id": row["stato"],
        })
        Lezione.objects.update_or_create(id=row["id"], defaults=defaults)

    ## Al ricarico completo, cancelliamo le lezioni e le sezioni non più presenti nella fonte (Excel)
    Lezione.objects.exclude(id__in=source_ids).delete()

    ## Cancello la tabella presente su db se presente e la ricreo (sarebbe un modo per creare una vista tramite codice)
    StrutturaLezione.objects.all().delete()
    StrutturaLezione.objects.bulk_create([StrutturaLezione(
        lezione_id=row["lezione_id"], ordine=row["ordine"], tipo_sezione=row["tipo_sezione"],
        contenuto=row["contenuto"], formato_web=row.get("formato_web", "testo"),
    ) for row in data.get("sezioni", [])])

    ## Stessa cosa per i quiz: cancelliamo e ricreiamo la tabella dei quiz e dei quesiti
    StrutturaQuiz.objects.all().delete()
    for row in data.get("quiz", []):
        quiz = StrutturaQuiz.objects.create(lezione_id=row["lezione_id"], modalita=row["modalita"], titolo=row["titolo"])
        # I quesiti finiscono nella tabella della propria modalita'.
        model = QuesitoGuidato if row["modalita"] == StrutturaQuiz.GUIDATO else QuesitoFinale
        model.objects.bulk_create([model(quiz=quiz, **question) for question in row.get("quesiti", [])])

    ## Cancelliamo le righe delle tabelle di lookup che non sono più presenti nella fonte (Excel)
    for key, model in LOOKUP_MODELS.items():
        source_codes = [str(item["code"]) for item in data["liste"][key]]
        model.objects.exclude(code__in=source_codes).delete()
    return {"lezioni": len(source_ids)}
